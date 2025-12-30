"""
Fund Analysis Excel Processor

A production-ready Python class to process fund analysis Excel files,
calculate monthly stock returns and contributions, and generate pivot table summaries.
"""

import logging
import pandas as pd
from pathlib import Path
from typing import Optional, Dict, Tuple
import numpy as np

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class FundAnalysisProcessor:
    """
    Processes fund analysis Excel files to calculate derived metrics and generate pivot tables.
    
    This class handles:
    - Loading Excel data with validation
    - Calculating Start Price, Monthly Stock Return%, Start wt%, and Stock Monthly Contribution %
    - Generating pivot tables for Company, Sector, and Market Cap analysis
    - Exporting results to Excel with multiple sheets
    """
    
    # Required input columns (A-K)
    REQUIRED_COLUMNS = [
        'Scheme Code',
        'Scheme Name',
        'Month',
        'Month End',
        'Instrument Name',
        'Holding (%)',
        'Instrument Sector',
        'Instrument SEBI Mcap',
        'Instrument SEBI Mcap Type',
        'NSE Symbol',
        'Price'
    ]
    
    # Calculated columns (L-O)
    CALCULATED_COLUMNS = [
        'Start Price',
        'Monthly Stock Return%',
        'Start wt%',
        'Stock Monthly Contribution %'
    ]
    
    def __init__(self, input_file_path: str, output_file_path: Optional[str] = None):
        """
        Initialize the FundAnalysisProcessor.
        
        Args:
            input_file_path: Path to the input Excel file
            output_file_path: Path for the output Excel file. If None, will be generated
                            from input_file_path with '_processed' suffix.
        
        Raises:
            ValueError: If input_file_path is empty or invalid
        """
        if not input_file_path:
            raise ValueError("Input file path cannot be empty")
        
        self.input_file_path = Path(input_file_path)
        if not self.input_file_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_file_path}")
        
        if output_file_path:
            self.output_file_path = Path(output_file_path)
        else:
            # Generate output path from input path
            self.output_file_path = self.input_file_path.parent / f"{self.input_file_path.stem}_processed.xlsx"
        
        self.df: Optional[pd.DataFrame] = None
        self.pivot_tables: Dict[str, pd.DataFrame] = {}
        
        logger.info(f"Initialized processor with input: {self.input_file_path}")
        logger.info(f"Output will be saved to: {self.output_file_path}")
    
    def load_data(self) -> pd.DataFrame:
        """
        Load data from Excel file with validation and error handling.
        
        Returns:
            DataFrame containing the loaded data
        
        Raises:
            FileNotFoundError: If the input file doesn't exist
            ValueError: If required columns are missing
            Exception: For other file reading errors
        """
        try:
            logger.info(f"Loading data from {self.input_file_path}")
            
            # Read Excel file
            self.df = pd.read_excel(self.input_file_path, engine='openpyxl')
            
            if self.df.empty:
                raise ValueError("Input Excel file is empty")
            
            logger.info(f"Loaded {len(self.df)} rows from Excel file")
            
            # Validate required columns
            self._validate_columns()
            
            # Ensure data is sorted by Instrument Name and Month
            self.df = self.df.sort_values(['Instrument Name', 'Month'], ascending=[True, True])
            self.df = self.df.reset_index(drop=True)
            
            logger.info("Data loaded and validated successfully")
            return self.df
            
        except FileNotFoundError:
            logger.error(f"File not found: {self.input_file_path}")
            raise
        except Exception as e:
            logger.error(f"Error loading data: {str(e)}")
            raise
    
    def _validate_columns(self) -> None:
        """
        Validate that all required columns exist in the dataframe.
        
        Raises:
            ValueError: If any required columns are missing
        """
        missing_columns = [col for col in self.REQUIRED_COLUMNS if col not in self.df.columns]
        
        if missing_columns:
            raise ValueError(
                f"Missing required columns: {missing_columns}. "
                f"Found columns: {list(self.df.columns)}"
            )
        
        logger.info("All required columns validated")
    
    def calculate_derived_columns(self) -> pd.DataFrame:
        """
        Calculate derived columns: Start Price, Monthly Stock Return%, Start wt%, 
        and Stock Monthly Contribution %.
        
        Returns:
            DataFrame with calculated columns added
        
        Raises:
            ValueError: If data hasn't been loaded yet
        """
        if self.df is None:
            raise ValueError("Data must be loaded before calculating derived columns. Call load_data() first.")
        
        logger.info("Calculating derived columns...")
        
        # Initialize calculated columns
        for col in self.CALCULATED_COLUMNS:
            if col not in self.df.columns:
                self.df[col] = np.nan
        
        # Calculate Start Price (L) - Previous month's Price for same Instrument Name
        self.df['Start Price'] = self.df.groupby('Instrument Name')['Price'].shift(1)
        
        # Calculate Monthly Stock Return% (M) - (Current Price / Start Price) - 1
        # Handle division by zero and missing values
        self.df['Monthly Stock Return%'] = np.where(
            (self.df['Start Price'].notna()) & (self.df['Start Price'] != 0),
            (self.df['Price'] / self.df['Start Price']) - 1,
            np.nan
        )
        
        # Calculate Start wt% (N) - Previous month's Holding (%) converted to decimal
        # Excel stores percentages as decimals (e.g., 3.06% stored as 0.0306)
        # If values are > 1, they're in percentage form and need conversion
        # If values are <= 1, they're already in decimal form
        holding_values = self.df['Holding (%)'].copy()
        
        # Check if values need conversion from percentage to decimal
        if holding_values.notna().any():
            sample_values = holding_values.dropna()
            if len(sample_values) > 0:
                max_value = sample_values.abs().max()
                # If max value > 1, assume values are in percentage form (e.g., 3.06 for 3.06%)
                # Convert to decimal (divide by 100)
                if max_value > 1:
                    holding_values = holding_values / 100
        
        # Store converted values temporarily in dataframe for groupby operation
        self.df['_holding_decimal_temp'] = holding_values
        
        # Get previous month's holding for same instrument
        self.df['Start wt%'] = self.df.groupby('Instrument Name')['_holding_decimal_temp'].shift(1)
        
        # Remove temporary column
        self.df = self.df.drop(columns=['_holding_decimal_temp'])
        
        # Fill NaN with 0 for first entry of each instrument
        self.df['Start wt%'] = self.df['Start wt%'].fillna(0)
        
        # Calculate Stock Monthly Contribution % (O) - Start wt% * Monthly Stock Return %
        self.df['Stock Monthly Contribution %'] = (
            self.df['Start wt%'] * self.df['Monthly Stock Return%']
        )
        
        # Fill NaN with 0 for contribution
        self.df['Stock Monthly Contribution %'] = self.df['Stock Monthly Contribution %'].fillna(0)
        
        logger.info("Derived columns calculated successfully")
        return self.df
    
    def create_pivot_tables(self) -> Dict[str, pd.DataFrame]:
        """
        Create pivot tables for Company, Sector, and Market Cap analysis.
        
        Returns:
            Dictionary containing pivot tables with keys: 'company', 'sector', 'market_cap'
        
        Raises:
            ValueError: If data hasn't been loaded or calculated yet
        """
        if self.df is None:
            raise ValueError("Data must be loaded before creating pivot tables. Call load_data() first.")
        
        if 'Stock Monthly Contribution %' not in self.df.columns:
            raise ValueError("Derived columns must be calculated first. Call calculate_derived_columns() first.")
        
        logger.info("Creating pivot tables...")
        
        # Company Pivot: Group by Instrument Name and Month End
        company_pivot = pd.pivot_table(
            self.df,
            values='Stock Monthly Contribution %',
            index='Instrument Name',
            columns='Month End',
            aggfunc='sum',
            fill_value=0,
            margins=True,
            margins_name='Grand Total'
        )
        
        # Sector Pivot: Group by Instrument Sector and Month End
        sector_pivot = pd.pivot_table(
            self.df,
            values='Stock Monthly Contribution %',
            index='Instrument Sector',
            columns='Month End',
            aggfunc='sum',
            fill_value=0,
            margins=True,
            margins_name='Grand Total'
        )
        
        # Market Cap Pivot: Group by Instrument SEBI Mcap Type and Month End
        market_cap_pivot = pd.pivot_table(
            self.df,
            values='Stock Monthly Contribution %',
            index='Instrument SEBI Mcap Type',
            columns='Month End',
            aggfunc='sum',
            fill_value=0,
            margins=True,
            margins_name='Grand Total'
        )
        
        self.pivot_tables = {
            'company': company_pivot,
            'sector': sector_pivot,
            'market_cap': market_cap_pivot
        }
        
        logger.info("Pivot tables created successfully")
        return self.pivot_tables
    
    def _format_percentage_columns(self) -> None:
        """
        Format percentage columns in the Excel output file.
        
        This method applies percentage formatting to columns that contain percentage values
        so they display correctly in Excel (e.g., 0.1199 displays as 11.99%).
        """
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(self.output_file_path)
            
            # Format percentage columns in Processed Data sheet
            if 'Processed Data' in workbook.sheetnames:
                ws = workbook['Processed Data']
                
                # Find column indices for percentage columns
                header_row = 1
                for col_idx, cell in enumerate(ws[header_row], start=1):
                    if cell.value in ['Monthly Stock Return%', 'Stock Monthly Contribution %', 'Holding (%)']:
                        # Format all data cells in this column as percentage
                        for row_idx in range(2, ws.max_row + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if cell.value is not None and isinstance(cell.value, (int, float)):
                                cell.number_format = '0.00%'
            
            # Format percentage columns in pivot tables
            for sheet_name in ['Company Pivot', 'Sector Pivot', 'Market Cap Pivot']:
                if sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    # Format all numeric cells as percentage (pivot tables contain contribution %)
                    for row in ws.iter_rows(min_row=2, min_col=2):
                        for cell in row:
                            if cell.value is not None and isinstance(cell.value, (int, float)):
                                cell.number_format = '0.00%'
            
            workbook.save(self.output_file_path)
            logger.info("Percentage columns formatted in Excel output")
            
        except Exception as e:
            logger.warning(f"Could not format percentage columns: {str(e)}. Values are stored as decimals.")
    
    def save_output(self) -> Path:
        """
        Save processed data and pivot tables to Excel file with multiple sheets.
        
        Returns:
            Path to the saved output file
        
        Raises:
            ValueError: If data or pivot tables haven't been created yet
            PermissionError: If unable to write to output file
        """
        if self.df is None:
            raise ValueError("No data to save. Call load_data() and calculate_derived_columns() first.")
        
        if not self.pivot_tables:
            raise ValueError("Pivot tables not created. Call create_pivot_tables() first.")
        
        try:
            logger.info(f"Saving output to {self.output_file_path}")
            
            # Create Excel writer
            with pd.ExcelWriter(self.output_file_path, engine='openpyxl') as writer:
                # Sheet 1: Processed data with all columns
                self.df.to_excel(writer, sheet_name='Processed Data', index=False)
                
                # Sheet 2: Company Pivot Table
                self.pivot_tables['company'].to_excel(writer, sheet_name='Company Pivot')
                
                # Sheet 3: Sector Pivot Table
                self.pivot_tables['sector'].to_excel(writer, sheet_name='Sector Pivot')
                
                # Sheet 4: Market Cap Pivot Table
                self.pivot_tables['market_cap'].to_excel(writer, sheet_name='Market Cap Pivot')
            
            # Format percentage columns in Excel
            self._format_percentage_columns()
            
            logger.info(f"Output saved successfully to {self.output_file_path}")
            return self.output_file_path
            
        except PermissionError:
            logger.error(f"Permission denied. Cannot write to {self.output_file_path}")
            raise
        except Exception as e:
            logger.error(f"Error saving output: {str(e)}")
            raise
    
    def process(self) -> Path:
        """
        Complete processing pipeline: load data, calculate columns, create pivots, and save output.
        
        Returns:
            Path to the saved output file
        """
        logger.info("Starting complete processing pipeline...")
        
        self.load_data()
        self.calculate_derived_columns()
        self.create_pivot_tables()
        output_path = self.save_output()
        
        logger.info("Processing completed successfully")
        return output_path


if __name__ == "__main__":
    """
    Example usage of FundAnalysisProcessor
    """
    import sys
    
    # Example: Process the test.xlsx file
    input_file = r"d:\car analysis\fund analysis\test.xlsx"
    
    try:
        # Initialize processor
        processor = FundAnalysisProcessor(
            input_file_path=input_file,
            output_file_path=None  # Will auto-generate output path
        )
        
        # Run complete processing pipeline
        output_file = processor.process()
        
        print(f"\n✓ Processing completed successfully!")
        print(f"✓ Output saved to: {output_file}")
        print(f"\nOutput contains 4 sheets:")
        print("  1. Processed Data - All original and calculated columns")
        print("  2. Company Pivot - Grouped by Instrument Name")
        print("  3. Sector Pivot - Grouped by Instrument Sector")
        print("  4. Market Cap Pivot - Grouped by Market Cap Type")
        
    except Exception as e:
        print(f"\n✗ Error during processing: {str(e)}", file=sys.stderr)
        sys.exit(1)

